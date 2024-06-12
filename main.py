import pandas as pd
import sqlite3
import re
from collections import Counter
from ctransformers import AutoModelForCausalLM
import openpyxl
from nltk.corpus import wordnet
import torch
from transformers import BertTokenizer, BertForSequenceClassification
from tqdm import tqdm
import os
import time

# ЭТАП 1 - ПОИСК СЛОВ ЭТАЛОНОВ
print('Начало 1 этапа.')
# Установить при первом запуске
# nltk.download('wordnet')

# Загружаем словарь английского языка
df = pd.read_excel('data/english_dictionary.xlsx')

# Определяем функцию для получения синонимов
def get_synonyms(word):
    synonyms = []
    synonyms.append(word)
    for syn in wordnet.synsets(word):
        for lemma in syn.lemmas():
            synonyms.append(lemma.name())

    return list(set(synonyms))

# Инициализируем начальную характеристику:
print('Введите стартовое слово:')
word_to_find_synonyms = input()
words_to_find = get_synonyms(word_to_find_synonyms)

if words_to_find:
    print(f"Синонимы для слова '{word_to_find_synonyms}': {', '.join(words_to_find)}")
else:
    print(f"Синонимы для слова '{word_to_find_synonyms}' не найдены.")


# Создание списка существительных
nouns_list = []

# Функция для поиск слов-эталонов
def collect_nouns(row):
    word = row['word']
    pos = row['pos']
    meaning = row.get('def', '')

    if pd.notna(pos) and 'n.' in pos:
        if isinstance(meaning, str) and any(word_to_find.lower() in meaning.lower() for word_to_find in words_to_find):
            nouns_list.append(word)

# Применяем функцию к каждой строке
df.apply(collect_nouns, axis=1)

# Создаем новый DataFrame из собранных существительных
result_df = pd.DataFrame({'Nouns': nouns_list})

# Удаляем дубликаты
result_df['Nouns'] = result_df['Nouns'].drop_duplicates()

# Сохраняем результаты этапа в таблицу
result_df.to_excel('data/standard_nouns.xlsx', index=False)

# Выводим краткий список имен-эталонов
print(result_df)

# Этап 2 - ПОИСК ПОТЕНЦИАЛЬНЫХ КЛАССИФИКАТОРОВ

print('Начало 2 этапа: поиск потенциальных классификаторов')

start_time = time.time()

# Подключаемся к базе данных
connection = sqlite3.connect('data/enwiki.db')
cursor = connection.cursor()

# Делаем запрос к БД
query = "SELECT SECTION_TEXT FROM ARTICLES ORDER BY ARTICLE_ID LIMIT 10000;"
cursor.execute(query)

with open('data/stop_words.txt', 'r') as file:
    stop_words = [line.strip() for line in file]

word_statistics = Counter()

# Список слов, для которых ищем соседей
df = pd.read_excel('data/standard_nouns.xlsx', sheet_name='Sheet1')

target_words = [str(word).strip() for word in df['Nouns'].dropna().tolist()]
target_words = [word.lower() for word in target_words]

try:
    while True:
        texts = cursor.fetchmany(100)
        if not texts:
            break

        for text_tuple in texts:
            text = text_tuple[0]
            words = re.findall(r'\b\w+\b', text.lower())

            for target_word in target_words:
                neighbors = []
                for i, word in enumerate(words):
                    if word == target_word:
                        context_window = 1
                        start_index = max(0, i - context_window)
                        end_index = min(len(words), i + context_window + 1)
                        context = words[start_index:end_index]
                        if target_word in context:
                            context.remove(target_word)
                        neighbors.extend(context)

                        for neighbor in context:
                            if neighbor not in stop_words:
                                word_statistics[neighbor] += 1

finally:
    wb = openpyxl.Workbook()
    ws = wb.active

    for row, (word, count) in enumerate(sorted(word_statistics.items(), key=lambda x: x[1], reverse=True), start=1):
        ws.cell(row=row, column=1, value=word)
        ws.cell(row=row, column=2, value=count)

    wb.save('data/potential_classifiers.xlsx')

    # Останавливаем таймер
    elapsed_time = time.time() - start_time
    print(f"\nПроцесс завершен. Окончательный результат сохранен в 'potential_classifiers.xlsx'. Время выполнения: {elapsed_time:.2f} секунд.")


# ЭТАП 3 - СЛОВАРНАЯ ПРОВЕРКА КЛАССИФИКАТОРОВ
print('Начало 3 этапа: проверка дефиниций классификаторов в словаре')

# Путь к таблице потенциальных классификаторов
potential_classifiers_excel = 'data/potential_classifiers.xlsx'
df = pd.read_excel('data/english_dictionary.xlsx').astype(str)
new_df = pd.read_excel(potential_classifiers_excel).astype(str)

# Создаем множество для хранения уникальных слов
result_words_set = set()

# Приводим слова для поиска к нижнему регистру
words_to_find_lower = [word.lower() for word in words_to_find]

# Цикл по каждому слову из потенциальных классификаторов
for new_word_to_find in new_df.iloc[:, 0]:
    new_word_to_find_lower = str(new_word_to_find).lower()

    # Поиск нового слова в датасете
    new_word_rows = df[df.iloc[:, 0].str.lower() == new_word_to_find_lower]

    if not new_word_rows.empty:
        for _, row in new_word_rows.iterrows():
            if any(word_to_find_lower in row['def'].lower() for word_to_find_lower in words_to_find_lower):
                result_words_set.add(row.iloc[0])

# Создаем новый DataFrame из уникальных слов
result_df = pd.DataFrame({'Result Words': list(result_words_set)})

# Сохраняем результаты в таблицу
result_df.to_excel(f'data/potential_classifiers_checked.xlsx', index=False)

# Выводим результаты
print("\nResult Words:")
print(result_df)

#Этап 4 - ФИНАЛЬНАЯ ПРОВЕРКА КЛАССИФИКАТОРОВ
print('Начало 4 этапа: проверка классификаторов с помощью LLM')
# Установите gpu_layers, которые нужно выгрузить на GPU. Подбирайте параметр в зависимости от вашего GPU.
llm = AutoModelForCausalLM.from_pretrained("TheBloke/Mistral-7B-Instruct-v0.1-GGUF",
                                           model_file="mistral-7b-instruct-v0.1.Q4_K_M.gguf", model_type="mistral",
                                           gpu_layers=30)

words_with_potential = list(result_words_set)

# Создаем новую таблицу Excel и лист
workbook = openpyxl.Workbook()
sheet = workbook.active

# Записываем заголовок столбца
sheet.cell(row=1, column=1, value=f"classifiers")

# Инициализируем счетчик строк
row_count = 2

for word in words_with_potential:
    response = llm(f"Check if the following word has the '{word_to_find_synonyms}' characteristic: '{word}'. Return '1' if it does, otherwise '0'.")
    print(word, response)
    print()

    if '1' in response.strip():
        sheet.cell(row=row_count, column=1, value=word)
        row_count += 1

# Сохраняем финальную таблицу классификаторов
workbook.save("data/final_classifiers.xlsx")


#Этап 5 - СОЗДАНИЕ ОБЩЕЙ ТАБЛИЦЫ КЛАССИФИКАТОРЫ + АБСТРАКТНЫЕ СУЩЕСТВИТЕЛЬНЫЕ
print('Начало 5 этапа: создание коллокаций классификатор + абстрактное существительное')

# Загружаем таблицы Excel
df1 = pd.read_excel("data/abstract_nouns.xlsx")
df2 = pd.read_excel("data/final_classifiers.xlsx")

# Создаем новую таблицу с двумя столбцами
df3 = pd.DataFrame({
    'abstract': df1['abstract'],
    'classifiers': df2['classifiers']
})

# Сохраняем новую таблицу
df3.to_excel("data/abstract_plus_classifiers.xlsx", index=False)


# Этап 6 - ПАРСИНГ ПРЕДЛОЖЕНИЙ КЛАССИФИКАТОР + АБСТРАКТНОЕ СУЩЕСТВИТЕЛЬНОЕ
print('Начало 6 этапа: парсинг предложений "классификатор + абстрактное существительное"')

# Загружаем списки слов из Excel файла
df = pd.read_excel('data/abstract_plus_classifiers.xlsx')
list1 = set(df['abstract'].dropna().str.lower().tolist())
list2 = set(df['classifiers'].dropna().str.lower().tolist())


# Функция для проверки условий
def check_conditions(sentence, list1, list2):
    words = sentence.split()
    for i, word in enumerate(words):
        if word in list1:
            for j in range(max(0, i - 4), min(len(words), i + 5)):
                if words[j] in list2 and words[j] != word:
                    return i, j
    return None


# Функция для обрезки предложения до 25 слов вокруг целевых слов
def trim_sentence(sentence, index1, index2):
    words = sentence.split()
    if len(words) <= 25:
        return ' '.join(words)

    # Находим центр между index1 и index2
    center = (index1 + index2) // 2

    # Лучше определить окно вокруг центра
    half_window = 12
    start = max(0, center - half_window)
    end = start + 25

    if end > len(words):
        end = len(words)
        start = max(0, end - 25)

    return ' '.join(words[start:end])


# Подключаемся к базе данных
connection = sqlite3.connect('data/enwiki.db')
cursor = connection.cursor()

# Делаем запрос
query = "SELECT SECTION_TEXT FROM ARTICLES ORDER BY ARTICLE_ID;"
cursor.execute(query)

# Пополняем файл sentences.csv
saved_file = 'data/sentences.csv'
batch_size = 500
saved_sentences = []
list1_words = []
list2_words = []

# Целевое количество предложений, которое требуется найти
target_count = 100000
total_counter = 0

try:
    while total_counter < target_count:
        texts = cursor.fetchmany(100)
        if not texts:
            break

        for text_tuple in texts:
            text = text_tuple[0]
            sentences = re.split(r'(?<=[.!?])\s+', text)
            for sentence in sentences:
                sentence = sentence.strip()
                if sentence:
                    sentence = sentence.lower()
                    result = check_conditions(sentence, list1, list2)
                    if result:
                        index1, index2 = result
                        trimmed_sentence = trim_sentence(sentence, index1, index2)
                        saved_sentences.append(trimmed_sentence)
                        list1_words.append(sentence.split()[index1])
                        list2_words.append(sentence.split()[index2])
                        total_counter += 1
                        if total_counter >= target_count:
                            break

                if len(saved_sentences) >= batch_size:
                    df = pd.DataFrame({
                        'sentences': saved_sentences,
                        'abstract': list1_words,
                        'classifiers': list2_words
                    })

                    if os.path.isfile(saved_file):
                        df.to_csv(saved_file, mode='a', header=False, index=False)
                    else:
                        df.to_csv(saved_file, index=False)

                    saved_sentences = []
                    list1_words = []
                    list2_words = []

                    print(f'Сохранено {total_counter} предложений')
            if total_counter >= target_count:
                break

finally:
    if saved_sentences:
        df = pd.DataFrame({
            'sentences': saved_sentences,
            'abstract': list1_words,
            'classifiers': list2_words
        })

        if os.path.isfile(saved_file):
            df.to_csv(saved_file, mode='a', header=False, index=False)
        else:
            df.to_csv(saved_file, index=False)

    print(f'Процесс завершен на {total_counter} предложениях. Результаты сохранены в файле:', saved_file)

#Этап 7 - ТЕСТИРОВАНИЕ ПРЕДЛОЖЕНИЙ
print('Начало 7 этапа: проверка коллокаций на метафорическую сочетаемость')

# Указываем путь к модели и токенизатору
model_path = "models/bert_tuned"
tokenizer_path = "models/tokenizer_tuned"

# Загружаем предобученную модель BERT и токенизатор
tokenizer = BertTokenizer.from_pretrained(tokenizer_path)
model = BertForSequenceClassification.from_pretrained(model_path)

# Указываем путь к таблице с предложениями
data_path = "data/sentences.csv"
# Указываем путь для сохранения результата
output_path = "data/result.xlsx"

# Загружаем данные из таблицы
df = pd.read_csv(data_path)

# Создаем столбец для меток (0 или 1)
df['Метафора'] = ''

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# Переводим модель в режим оценки
model.eval()

# Классифицируем каждое предложение и ставим метку в столбец 'Метафора'
with torch.no_grad():
    for i, row in tqdm(df.iterrows(), total=len(df), desc="Classifying"):
        phrase = row['sentences']
        inputs = tokenizer(phrase, return_tensors="pt", truncation=True, padding=True).to(device)
        outputs = model(**inputs)
        logits = outputs.logits
        predictions = torch.argmax(logits, dim=1).item()
        df.at[i, 'Метафора'] = predictions

# Сохраняем результат в новый файл
df.to_excel(output_path, index=False)

# Этап 8 - ПОДСЧЕТ ПРОЦЕНТОВ
print('Начало 8 этапа: подсчет результатов')
def calculate_statistics(df, column_name):
    # Создаем новый DataFrame для хранения результатов
    result_df = pd.DataFrame(columns=['Слово', 'Количество', 'Количество_1', 'Количество_0', 'Процент_1'])

    # Получаем уникальные слова
    unique_words = df[column_name].unique()

    # Для каждого уникального слова считаем количество и количество значений 1 и 0 из столбца 'Метафора'
    for word in unique_words:
        total_count = df[df[column_name] == word].shape[0]
        count_1 = df[(df[column_name] == word) & (df['Метафора'] == 1)].shape[0]
        count_0 = df[(df[column_name] == word) & (df['Метафора'] == 0)].shape[0]
        percent_1 = (count_1 / total_count) * 100 if total_count > 0 else 0

        # Добавляем результаты в новый DataFrame
        result_df = pd.concat([result_df, pd.DataFrame({'Слово': [word], 'Количество': [total_count],
                                                         'Количество_1': [count_1], 'Количество_0': [count_0],
                                                         'Процент_1': [percent_1]})], ignore_index=True)

    # Сортируем по столбцу 'Процент_1'
    result_df = result_df.sort_values(by='Процент_1', ascending=False)

    # Сохраняем результат в новый файл
    result_df.to_excel(f"data/Результаты_{word_to_find_synonyms}_{column_name}.xlsx", index=False)
    print(f"Таблица с результатами для столбца '{column_name}' сохранена.")

# Путь к файлу с данными
data_path = "data/result.xlsx"

# Загружаем данные из таблицы
df = pd.read_excel(data_path)

# Вызываем функцию дважды - для 'abstract' и 'Result Words'
calculate_statistics(df, 'abstract')
calculate_statistics(df, 'classifiers')